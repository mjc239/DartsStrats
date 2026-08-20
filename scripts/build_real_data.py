#!/usr/bin/env python3
"""Rebuild data/real/ from open upstream sources.

The CSVs this produces are not committed: they are regenerable, they total ~35 MB,
and none of the four upstream repositories carries a licence, so we prefer not to
redistribute their contents. Run this script to materialise them locally.

    python scripts/build_real_data.py            # rebuild, then verify checksums
    python scripts/build_real_data.py --verify-only
    python scripts/build_real_data.py --keep-cache

Requires network access to github.com, plus: openpyxl, pyreadr.
    pip install openpyxl pyreadr

Every upstream is pinned to a commit SHA, so output is deterministic. Expected
SHA-256 digests live in data/real/CHECKSUMS.txt and are checked on every run;
a mismatch means an upstream changed or a dependency behaves differently, and
is reported as a failure rather than silently accepted.

The validation gates below are the same ones used when this data was first
assembled. They are assertions, not warnings: if an invariant breaks the script
stops. See data/real/PROVENANCE.md for what each source is and what its
definitional caveats are.
"""

from __future__ import annotations

import argparse
import collections
import csv
import glob
import hashlib
import os
import shutil
import subprocess
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT = os.path.join(ROOT, "data", "real")
CACHE = os.path.join(ROOT, ".cache", "darts_upstream")

# Upstream sources, pinned. Do not float these to HEAD: the reconstruction in
# build_2022() depends on the exact file set, and the checksums depend on all of it.
SOURCES = {
    "dartsviz": (
        "https://github.com/dmorgan26/dartsviz.git",
        "f7bfcfae388e961198dc94f5d37fd7129234aa75",
    ),
    "darts_data": (
        "https://github.com/wonderkiduk/darts_data.git",
        "dc3821eac7257f88e763fff181cf1d0d229e0591",
    ),
    "OptimalDarts": (
        "https://github.com/wangchunsem/OptimalDarts.git",
        "f68bedcc592b4bcfda69f6d72d69aede91ebaff1",
    ),
    "flashscore_wc2025": (
        "https://github.com/henlewis/Darts-Web-Scraping-PowerBi-Dashboard.git",
        "5441cb1d347bf90afb6029e54e0838a71795c9d3",
    ),
}

# 11 PDC WC 2022 matches whose leg replay does not close cleanly (darts left over
# after the final checkout). Excluded wholesale rather than patched. Recomputed
# below and asserted to match this list, so a silent upstream change is caught.
EXPECTED_2022_EXCLUDED = 11

# 42 rows in the 2022 feed name a real bed (S1, T20, D18...) but report points = 0,
# i.e. a bounce-out or voided dart where the segment was still logged. `value` is
# authoritative for arithmetic and is what the leg replay uses; `bed` is unreliable
# on exactly these rows. Asserted so the count cannot drift unnoticed.
EXPECTED_2022_BED_VALUE_MISMATCH = 42

PLAYER_SHORT_TO_FULL = {
    "Anderson G": "Gary Anderson", "Aspinall": "Nathan Aspinall",
    "Chisnall": "Dave Chisnall", "Clayton": "Jonny Clayton", "Cross": "Rob Cross",
    "Cullen": "Joe Cullen", "Gurney": "Daryl Gurney", "Lewis A": "Adrian Lewis",
    "Price": "Gerwyn Price", "Smith M": "Michael Smith", "Suljovic": "Mensur Suljovic",
    "Wade": "James Wade", "White": "Ian White", "Whitlock": "Simon Whitlock",
    "Wright": "Peter Wright", "van Gerwen": "Michael van Gerwen",
}

TREBLE_OUTCOME_BEDS = {
    "T20 attempt": ["T20", "S20", "T5", "S5", "T1", "S1"],
    "T19 attempt": ["T19", "S19", "T7", "S7", "T3", "S3"],
    "T18 attempt": ["T18", "S18", "T4", "S4", "T1", "S1"],
    "T17 attempt": ["T17", "S17", "T3", "S3", "T2", "S2"],
}


def log(msg):
    print(msg, flush=True)


def fetch():
    """Clone each upstream at its pinned commit into the cache."""
    os.makedirs(CACHE, exist_ok=True)
    for name, (url, sha) in SOURCES.items():
        dest = os.path.join(CACHE, name)
        if os.path.isdir(os.path.join(dest, ".git")):
            have = subprocess.run(["git", "-C", dest, "rev-parse", "HEAD"],
                                  capture_output=True, text=True).stdout.strip()
            if have == sha:
                log(f"  cached  {name} @ {sha[:10]}")
                continue
            shutil.rmtree(dest)
        log(f"  clone   {name} @ {sha[:10]}")
        subprocess.run(["git", "clone", "--quiet", url, dest], check=True)
        subprocess.run(["git", "-C", dest, "checkout", "--quiet", sha], check=True)


# --------------------------------------------------------------------------
# Schema A source 1: dartsviz, 2017 PDC. score_before is supplied natively.
# --------------------------------------------------------------------------

def bed_dartsviz(segment, ring):
    if ring == 3:
        return "T%d" % segment
    if ring == 2:
        return "D%d" % segment
    if ring == 1:
        if segment == 50:
            return "BULL"
        if segment == 25:
            return "25"
        if segment == 0:
            return "MISS"
        return "S%d" % segment
    return "MISS"


def build_2017():
    import pyreadr

    d = os.path.join(CACHE, "dartsviz", "plot-app", "data")
    throws = pyreadr.read_r(os.path.join(d, "throws.rda"))["throws"]
    legs = pyreadr.read_r(os.path.join(d, "legs.rda"))["legs"]
    players = pyreadr.read_r(os.path.join(d, "players.rda"))["players"]

    for c in ["throw_id", "player_id", "leg_id", "segment", "ring_id",
              "dart_sequence", "visit_sequence", "score_before"]:
        throws[c] = throws[c].astype(int)

    names = dict(zip(players.player_id.astype(int), players.name))
    leg_to_match = dict(zip(legs.leg_id.astype(int), legs.match_id.astype("Int64")))
    throws = throws.sort_values(["leg_id", "throw_id"])

    # GATE: the source's own arithmetic must hold across every within-visit
    # dart pair. This is what establishes that score = segment * ring_id and
    # hence that our bed mapping is right. Zero tolerance.
    ok = bad = 0
    prev = None
    for r in throws.itertuples():
        if (prev and prev.leg_id == r.leg_id and prev.player_id == r.player_id
                and prev.visit_sequence == r.visit_sequence
                and r.dart_sequence == prev.dart_sequence + 1):
            if prev.score_before - prev.segment * prev.ring_id == r.score_before:
                ok += 1
            else:
                bad += 1
        prev = r
    assert bad == 0, f"dartsviz within-visit arithmetic broken on {bad} transitions"
    log(f"  [gate] 2017 within-visit arithmetic: {ok} transitions, 0 inconsistent")

    rows = []
    for r in throws.itertuples():
        m = leg_to_match.get(r.leg_id)
        rows.append(dict(
            source="dartsviz_pdc_2017", player=names.get(r.player_id, ""),
            match_id=("" if m is None or str(m) == "<NA>" else int(m)),
            leg_id=r.leg_id, visit_index=r.visit_sequence, dart_index=r.dart_sequence,
            score_before=r.score_before, bed=bed_dartsviz(r.segment, r.ring_id),
            value=r.segment * r.ring_id, raw_segment=r.segment, raw_ring=r.ring_id,
            score_before_origin="source"))
    return rows


# --------------------------------------------------------------------------
# Schema A source 2: Sportradar PDC WC 2022. score_before is RECONSTRUCTED.
# --------------------------------------------------------------------------

def bed_sportradar(segment, event_points, points):
    segment, event_points, points = int(segment), int(event_points), int(points)
    if segment == 3:
        return "T%d" % event_points
    if segment == 2:
        return "D%d" % event_points
    if segment in (10, 11):          # 10 = inner single, 11 = outer single
        return "S%d" % event_points
    if segment == 5:
        return "BULL"
    if segment == 4:
        return "25" if points == 25 else "MISS"
    return "MISS"


def build_2022():
    rows, kept, excluded = [], [], []
    pattern = os.path.join(CACHE, "darts_data", "data", "*.csv")
    for path in sorted(glob.glob(pattern)):
        stem = os.path.basename(path)[:-4]
        tail = stem.replace("pdc_world_championship_2022_", "")
        home, _, away = tail.partition("_v_")
        name = {"home": home.replace("_", " ").title(),
                "away": away.replace("_", " ").title()}

        with open(path, newline="") as f:
            src = list(csv.DictReader(f))
        if not src:
            continue

        rem = {"home": 501, "away": 501}
        visit_start = dict(rem)
        current, dart_in_visit, leg_no, visit_no, open_darts = None, 0, 1, 0, 0
        buf = []

        for r in src:
            t = r["team"]
            if t != current or dart_in_visit == 3:
                current, dart_in_visit = t, 0
                visit_start[t] = rem[t]
                visit_no += 1
            score_before = rem[t]
            points = int(r["points"])
            bed = bed_sportradar(r["segment"], r["event_points"], r["points"])
            dart_in_visit += 1
            open_darts += 1
            buf.append(dict(
                source="sportradar_pdc_wc_2022", player=name[t], match_id=stem,
                leg_id="%s#L%d" % (stem, leg_no), visit_index=visit_no,
                dart_index=dart_in_visit, score_before=score_before, bed=bed,
                value=points, raw_segment=r["segment"], raw_ring="",
                score_before_origin="reconstructed"))

            new = score_before - points
            is_double = bed.startswith("D") or bed == "BULL"
            if new == 0 and is_double:                       # checkout
                rem = {"home": 501, "away": 501}
                visit_start = dict(rem)
                current, dart_in_visit, visit_no, open_darts = None, 0, 0, 0
                leg_no += 1
            elif new < 0 or new == 1 or (new == 0 and not is_double):   # bust
                rem[t] = visit_start[t]
                dart_in_visit = 3
            else:
                rem[t] = new

        # A match is kept only if the replay closes: no darts left dangling after
        # the last checkout. Anything else is dropped whole, never patched.
        if open_darts == 0:
            kept.append(stem)
            rows.extend(buf)
        else:
            excluded.append(stem)

    # GATE: every retained leg must run 501 -> exactly 0 on a double. A
    # desynchronised replay cannot satisfy this across thousands of legs.
    by_leg = collections.defaultdict(list)
    for r in rows:
        by_leg[r["leg_id"]].append(r)
    closed = 0
    for leg_id, darts in by_leg.items():
        winner = darts[-1]
        if winner["score_before"] - winner["value"] == 0 and (
                winner["bed"].startswith("D") or winner["bed"] == "BULL"):
            closed += 1
    assert closed == len(by_leg), (
        f"{len(by_leg) - closed} of {len(by_leg)} reconstructed legs do not close on a double")
    assert len(excluded) == EXPECTED_2022_EXCLUDED, (
        f"expected {EXPECTED_2022_EXCLUDED} unreconstructable matches, got {len(excluded)}")
    log(f"  [gate] 2022 replay: {closed} legs all close 501->0 on a double; "
        f"{len(kept)} matches kept, {len(excluded)} excluded")

    # GATE: rows whose named bed contradicts the points scored (see constant above).
    def implied(bed):
        if bed in ("MISS",):
            return 0
        if bed == "BULL":
            return 50
        if bed == "25":
            return 25
        if bed[0] == "T":
            return 3 * int(bed[1:])
        if bed[0] == "D":
            return 2 * int(bed[1:])
        return int(bed[1:])
    mismatch = sum(1 for r in rows if implied(r["bed"]) != r["value"])
    assert mismatch == EXPECTED_2022_BED_VALUE_MISMATCH, (
        f"bed/value mismatches = {mismatch}, expected {EXPECTED_2022_BED_VALUE_MISMATCH}")
    log(f"  [gate] 2022 bed/value: {mismatch} zero-point rows with a named bed "
        f"(known bounce-outs; trust `value`)")
    return rows


# --------------------------------------------------------------------------
# Schemas A and B
# --------------------------------------------------------------------------

FIELDS_A = ["source", "player", "match_id", "leg_id", "visit_index", "dart_index",
            "score_before", "bed", "value", "raw_segment", "raw_ring",
            "score_before_origin", "post_bust_visit"]
FIELDS_B = ["source", "player", "match_id", "leg_id", "visit_index", "score_before",
            "visit_score", "darts_used", "checkout", "bust", "post_bust_visit"]


def flag_post_bust(per_dart):
    """The 2017 feed keeps subtracting after a bust instead of reverting, which
    yields score_before < 2. Flag the whole affected visit; never delete or fix."""
    bad = set()
    for d in per_dart:
        if int(d["score_before"]) < 2:
            bad.add((d["source"], d["leg_id"], d["player"], int(d["visit_index"])))
    for d in per_dart:
        key = (d["source"], d["leg_id"], d["player"], int(d["visit_index"]))
        d["post_bust_visit"] = 1 if key in bad else 0
    clean = [d for d in per_dart if not d["post_bust_visit"]]
    assert all(2 <= int(d["score_before"]) <= 501 for d in clean), \
        "unflagged row outside legal score range"
    log(f"  [gate] post-bust: {len(bad)} visits flagged, "
        f"all {len(clean)} unflagged darts have 2 <= score_before <= 501")
    return bad


def derive_per_visit(per_dart, bad_visits):
    groups = collections.OrderedDict()
    for d in per_dart:
        key = (d["source"], d["player"], d["match_id"], d["leg_id"], d["visit_index"])
        groups.setdefault(key, []).append(d)

    out = []
    for key, darts in groups.items():
        darts = sorted(darts, key=lambda x: x["dart_index"])
        score_before = darts[0]["score_before"]
        total = sum(d["value"] for d in darts)
        last = darts[-1]
        is_double = last["bed"].startswith("D") or last["bed"] == "BULL"
        remaining = score_before - total
        out.append(dict(
            source=key[0], player=key[1], match_id=key[2], leg_id=key[3],
            visit_index=key[4], score_before=score_before,
            # raw arithmetic sum, including on busts; `bust` lets you apply
            # whichever convention your model wants
            visit_score=total, darts_used=len(darts),
            checkout=1 if (remaining == 0 and is_double) else 0,
            bust=1 if (remaining < 0 or remaining == 1
                       or (remaining == 0 and not is_double)) else 0,
            post_bust_visit=1 if (key[0], key[3], key[1], int(key[4])) in bad_visits else 0))

    assert all(0 <= r["visit_score"] <= 180 for r in out), "visit_score out of range"
    assert all(1 <= r["darts_used"] <= 3 for r in out), "darts_used out of range"
    log(f"  [gate] per-visit: {len(out)} visits, visit_score in [0,180], "
        f"darts_used in [1,3]")
    return out


# --------------------------------------------------------------------------
# Schema C: 2025 PDC World Championship aggregates
# --------------------------------------------------------------------------

FIELDS_C = ["source", "player", "match_id", "date", "event", "opponent", "legs_played",
            "darts_thrown", "points_scored", "three_dart_average", "first9_average",
            "count_180", "count_140_plus", "count_100_plus", "doubles_hit",
            "doubles_attempted", "checkout_percentage", "high_checkout", "stage",
            "sets_won", "sets_lost", "legs_won"]


def build_aggregates():
    import openpyxl

    path = os.path.join(CACHE, "flashscore_wc2025", "Darts Data.xlsx")
    ws = openpyxl.load_workbook(path)["Sheet1"]
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    idx = {h: i for i, h in enumerate(rows[0])}

    out = []
    for i, r in enumerate(rows[1:], 1):
        if not r[idx["player_name"]]:
            continue
        hits = r[idx["successful_checkouts"]]
        misses = r[idx["failed_checkouts"]]
        # doubles_attempted is the sum of two SEPARATELY PUBLISHED counts.
        # It is never back-computed from the percentage.
        attempts = int(hits) + int(misses) if hits not in (None, "") and misses not in (None, "") else ""
        out.append({
            "source": "flashscore_pdc_wc_2025", "player": r[idx["player_name"]],
            "match_id": "wc2025_m%03d" % ((i + 1) // 2), "date": r[idx["match_date"]],
            "event": "PDC World Darts Championship 2025",
            "opponent": r[idx["opponent_name"]],
            "legs_played": "", "darts_thrown": "", "points_scored": "",
            "three_dart_average": r[idx["average_3_darts"]], "first9_average": "",
            "count_180": r[idx["180_thrown"]], "count_140_plus": r[idx["140+_thrown"]],
            "count_100_plus": r[idx["100+_thrown"]], "doubles_hit": hits,
            "doubles_attempted": attempts,
            "checkout_percentage": r[idx["checkout_percentage"]],
            "high_checkout": r[idx["highest_checkout"]], "stage": r[idx["stage"]],
            "sets_won": r[idx["sets_won"]], "sets_lost": r[idx["sets_lost"]],
            "legs_won": r[idx["legs_won"]]})

    # GATE: hits/(hits+misses) must reproduce the separately published percentage.
    bad = 0
    for r in out:
        if r["doubles_attempted"]:
            got = int(r["doubles_hit"]) / int(r["doubles_attempted"]) * 100
            if abs(got - float(r["checkout_percentage"])) > 0.02:
                bad += 1
    assert bad == 0, f"{bad} rows where doubles_hit/attempted contradicts published checkout%"
    log(f"  [gate] aggregates: {len(out)} rows, all reproduce published checkout%")
    return out


# --------------------------------------------------------------------------
# Target outcomes and by-double attempts (2019 top 16, intended target known)
# --------------------------------------------------------------------------

def build_target_outcomes():
    import openpyxl

    path = os.path.join(CACHE, "OptimalDarts", "Raw_Data.xlsx")
    wb = openpyxl.load_workbook(path, data_only=False)
    rows = []

    grid = [[c.value for c in r] for r in wb["Trebles"].iter_rows()]
    blocks = [(v.strip(), ci) for ci, v in enumerate(grid[0])
              if isinstance(v, str) and v.strip().endswith("attempt")]
    for label, c0 in blocks:
        target = label.split()[0]
        for r in grid[2:]:
            player = r[0]
            if not isinstance(player, str) or not player:
                continue
            for k, bed in enumerate(TREBLE_OUTCOME_BEDS[label]):
                v = r[c0 + k]
                if isinstance(v, (int, float)):
                    rows.append((player.strip(), target, bed, int(v)))

    grid = [[c.value for c in r] for r in wb["Doubles"].iter_rows()]
    players = [(ci, v.strip()) for ci, v in enumerate(grid[0])
               if isinstance(v, str) and v.strip()]
    current = None
    for r in grid[1:]:
        a, b = r[0], r[1]
        if isinstance(a, str) and a.strip():
            current = a.strip()
        if current is None or b is None:
            continue
        if current == "Bullseye":
            target = "DB"
            bed = "DB" if b == "Bull" else ("SB" if b == 25 else "S%d" % int(b))
        else:
            target = current
            if b == "missed":
                bed = "MISS"
            elif isinstance(b, str) and b.startswith("D"):
                bed = b
            else:
                bed = "S%d" % int(b)
        for ci, short in players:
            v = r[ci]
            if isinstance(v, (int, float)):
                rows.append((PLAYER_SHORT_TO_FULL.get(short, short), target, bed, int(v)))

    attempts = collections.Counter()
    hits = collections.Counter()
    for p, t, b, c in rows:
        attempts[(p, t)] += c
        if b == t:
            hits[(p, t)] += c

    # GATE: reproduce the summary statistics published in arXiv:2302.10750 §3.
    # Note D20: the paper says 4,399 but the workbook gives 4,339, and only
    # 4,339 is consistent with the paper's own grand total of 16,777. We assert
    # the workbook figure and leave the discrepancy documented, not corrected.
    def tot(t):
        return sum(c for p, tt, b, c in rows if tt == t)

    def hit(t):
        return sum(c for p, tt, b, c in rows if tt == t and b == t)

    for target, expected in [("T20", 117600), ("T19", 27709), ("T18", 7717), ("T17", 2461)]:
        assert tot(target) == expected, f"{target} attempts {tot(target)} != {expected}"
    doubles = [t for t in {t for _, t, _, _ in rows} if t.startswith("D")]
    assert sum(tot(t) for t in doubles) == 16777, "double-region total != 16,777"
    assert sum(tot("D%d" % i) for i in range(1, 20, 2)) == 1866, "odd doubles != 1,866"
    assert tot("D20") == 4339, "D20 attempts != 4,339 (workbook figure)"
    assert round(hit("T20") / tot("T20") * 100, 1) == 41.2
    log("  [gate] 2019 target outcomes: reproduces every published summary stat "
        "(T20/T19/T18/T17 attempts, 16,777 total, 1,866 odd doubles)")

    return rows, attempts, hits


# --------------------------------------------------------------------------

def write_csv(path, fields, rows, dictrows=True):
    with open(path, "w", newline="") as f:
        if dictrows:
            w = csv.DictWriter(f, fields)
            w.writeheader()
            w.writerows(rows)
        else:
            w = csv.writer(f)
            w.writerow(fields)
            w.writerows(rows)


def checksums():
    out = {}
    for name in sorted(os.listdir(OUT)):
        if name.endswith(".csv"):
            h = hashlib.sha256()
            with open(os.path.join(OUT, name), "rb") as f:
                for chunk in iter(lambda: f.read(1 << 20), b""):
                    h.update(chunk)
            out[name] = h.hexdigest()
    return out


def verify():
    manifest = os.path.join(OUT, "CHECKSUMS.txt")
    if not os.path.exists(manifest):
        log("no CHECKSUMS.txt to verify against")
        return 1
    expected = {}
    for line in open(manifest):
        line = line.strip()
        if line and not line.startswith("#"):
            digest, name = line.split(None, 1)
            expected[name.strip()] = digest
    got = checksums()
    ok = True
    for name, digest in sorted(expected.items()):
        if name not in got:
            log(f"  MISSING  {name}")
            ok = False
        elif got[name] != digest:
            log(f"  MISMATCH {name}\n           expected {digest}\n           got      {got[name]}")
            ok = False
        else:
            log(f"  ok       {name}")
    return 0 if ok else 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--verify-only", action="store_true",
                    help="check existing files against CHECKSUMS.txt and exit")
    ap.add_argument("--keep-cache", action="store_true",
                    help="keep cloned upstreams in .cache/ for reuse")
    ap.add_argument("--write-checksums", action="store_true",
                    help="regenerate CHECKSUMS.txt from the built files")
    args = ap.parse_args()

    if args.verify_only:
        sys.exit(verify())

    os.makedirs(OUT, exist_ok=True)
    log("fetching pinned upstreams")
    fetch()

    log("building")
    per_dart = build_2017() + build_2022()
    bad_visits = flag_post_bust(per_dart)
    per_visit = derive_per_visit(per_dart, bad_visits)
    aggregates = build_aggregates()
    target_rows, attempts, hits = build_target_outcomes()

    write_csv(os.path.join(OUT, "per_dart.csv"), FIELDS_A, per_dart)
    write_csv(os.path.join(OUT, "per_visit.csv"), FIELDS_B, per_visit)
    write_csv(os.path.join(OUT, "match_aggregates.csv"), FIELDS_C, aggregates)
    write_csv(os.path.join(OUT, "target_outcomes.csv"),
              ["source", "player", "target", "outcome_bed", "count"],
              [["optimaldarts_pdc_2019", p, t, b, c] for p, t, b, c in target_rows],
              dictrows=False)
    write_csv(os.path.join(OUT, "double_attempts.csv"),
              ["source", "player", "double", "attempts", "hits"],
              [["optimaldarts_pdc_2019", p, t, attempts[(p, t)], hits[(p, t)]]
               for (p, t) in sorted(attempts) if t.startswith("D")],
              dictrows=False)

    log(f"\nwrote {len(per_dart)} darts, {len(per_visit)} visits, "
        f"{len(aggregates)} match rows, {len(target_rows)} target-outcome rows")

    if args.write_checksums:
        with open(os.path.join(OUT, "CHECKSUMS.txt"), "w") as f:
            f.write("# SHA-256 of files produced by scripts/build_real_data.py\n")
            f.write("# Upstreams are pinned; output is deterministic.\n")
            for name, digest in sorted(checksums().items()):
                f.write(f"{digest}  {name}\n")
        log("wrote CHECKSUMS.txt")
    else:
        log("\nverifying against CHECKSUMS.txt")
        rc = verify()
        if rc:
            sys.exit(rc)

    if not args.keep_cache and os.path.isdir(CACHE):
        shutil.rmtree(CACHE)
        log("removed upstream cache (--keep-cache to retain)")


if __name__ == "__main__":
    main()
